from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.db import get_conn
from app.services.snapshot_service import SnapshotService
from app.services.utils import now_iso, src_hash


class UpdateService:
    def __init__(self) -> None:
        self.snapshots = SnapshotService()

    def update_dev_from_directory(
        self,
        source_dir: str,
        lang: str,
        version_tag: str,
        parent_snapshot_id: int | None,
        target_col_index: int = 3,
    ) -> int:
        snapshot_id = self.snapshots.create_snapshot(
            branch="dev",
            parent_snapshot_id=parent_snapshot_id,
            action_type="update_dev",
            meta={"source_dir": source_dir, "lang": lang, "version_tag": version_tag},
        )
        if parent_snapshot_id:
            self.snapshots.copy_items(parent_snapshot_id, snapshot_id)

        root = Path(source_dir)
        files = [p for p in root.rglob("*.xlsx") if not p.name.startswith("~$")]

        for file in files:
            wb = load_workbook(file)
            for sheet in wb.worksheets:
                for row_idx in range(2, sheet.max_row + 1):
                    key = str(sheet.cell(row=row_idx, column=1).value or "").strip()
                    if not key:
                        continue
                    src = str(sheet.cell(row=row_idx, column=2).value or "")
                    tgt = str(sheet.cell(row=row_idx, column=target_col_index).value or "")
                    entry_id = self._upsert_entry(key, src, version_tag)
                    self._upsert_translation(entry_id, lang, tgt)
                    self.snapshots.set_item(snapshot_id, key, entry_id, src_hash(src))
        return snapshot_id

    def _upsert_entry(self, key: str, src: str, version_tag: str) -> int:
        src_digest = src_hash(src)
        with get_conn() as conn:
            row = conn.execute(
                "SELECT entry_id FROM entries WHERE key = ? AND src_hash = ? ORDER BY entry_id DESC LIMIT 1",
                (key, src_digest),
            ).fetchone()
            if row:
                conn.execute(
                    "UPDATE entries SET version_tag = ? WHERE entry_id = ?",
                    (version_tag, row["entry_id"]),
                )
                return int(row["entry_id"])
            cur = conn.execute(
                """
                INSERT INTO entries(key, src, src_hash, version_tag, meta_json)
                VALUES (?, ?, ?, ?, '{}')
                """,
                (key, src, src_digest, version_tag),
            )
            return int(cur.lastrowid)

    def _upsert_translation(self, entry_id: int, lang: str, tgt: str) -> None:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO translations(entry_id, lang, target_text, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(entry_id, lang)
                DO UPDATE SET target_text=excluded.target_text, updated_at=excluded.updated_at
                """,
                (entry_id, lang, tgt, now_iso()),
            )

    def update_release_active_single(self, release_snapshot_id: int, key: str, lang: str, target_text: str) -> int:
        base = self.snapshots.get_snapshot_items(release_snapshot_id)
        if key not in base:
            raise KeyError(f"key not found in release snapshot: {key}")
        new_snapshot = self.snapshots.create_snapshot(
            "release",
            "active_single",
            parent_snapshot_id=release_snapshot_id,
            meta={"key": key, "lang": lang},
        )
        self.snapshots.copy_items(release_snapshot_id, new_snapshot)
        self._upsert_translation(int(base[key]["entry_id"]), lang, target_text)
        return new_snapshot


    def update_release_passive_single(
        self,
        release_snapshot_id: int,
        key: str,
        src: str,
        targets_by_lang: dict[str, str],
        version_tag: str,
    ) -> int:
        new_snapshot = self.snapshots.create_snapshot(
            "release",
            "passive_single",
            parent_snapshot_id=release_snapshot_id,
            meta={"key": key},
        )
        self.snapshots.copy_items(release_snapshot_id, new_snapshot)
        entry_id = self._upsert_entry(key, src, version_tag)
        for lang, tgt in targets_by_lang.items():
            self._upsert_translation(entry_id, lang, tgt)
        self.snapshots.set_item(new_snapshot, key, entry_id, src_hash(src))
        return new_snapshot
