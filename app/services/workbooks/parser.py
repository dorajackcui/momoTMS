from __future__ import annotations

from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from app.services.project.service import DEFAULT_PROJECT_ID, ProjectService
from app.services.shared.io import normalize_content_value, normalize_non_content_value
from app.services.workbooks.models import (
    WorkbookPrecheck,
    WorkbookRow,
    WorkbookSheetPreview,
    WorkbookWorkflowContext,
)


class WorkbookParser:
    SAMPLE_ROW_LIMIT = 50

    def __init__(self, projects: ProjectService | None = None) -> None:
        self.projects = projects or ProjectService()

    def precheck_directory(
        self,
        input_dir: str | Path,
        project_id: int = DEFAULT_PROJECT_ID,
        context: WorkbookWorkflowContext | None = None,
    ) -> WorkbookPrecheck:
        context = context or WorkbookWorkflowContext(workflow_kind="create_branch")
        base = Path(input_dir)
        files = self._list_workbooks(base)
        sheet_previews: list[WorkbookSheetPreview] = []
        all_missing: list[str] = []
        sampled_issue_count = 0

        for file_path in files:
            rel_path = self._relative_path(base, file_path)
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    row_iter = sheet.iter_rows(values_only=True)
                    headers = list(next(row_iter, ()))
                    mapping = self._resolve_mapping(headers, project_id, context, strict=False)
                    missing = list(mapping["missing_required_headers"])
                    all_missing.extend(header for header in missing if header not in all_missing)
                    sampled_issue_count += self._sample_issue_count(row_iter, mapping, context)
                    sheet_previews.append(
                        WorkbookSheetPreview(
                            sheet_key=f"{rel_path}::{sheet.title}",
                            file_path=rel_path,
                            sheet_name=sheet.title,
                            available_headers=list(mapping["available_headers"]),
                            missing_required_headers=missing,
                            sampled_issue_count=sampled_issue_count,
                        )
                    )
            finally:
                workbook.close()

        return WorkbookPrecheck(
            file_count=len(files),
            sheet_count=len(sheet_previews),
            missing_required_headers=all_missing,
            sampled_issue_count=sampled_issue_count,
            sheet_previews=sheet_previews,
        )

    def iter_rows(
        self,
        input_dir: str | Path,
        project_id: int = DEFAULT_PROJECT_ID,
        context: WorkbookWorkflowContext | None = None,
    ) -> Iterator[WorkbookRow]:
        context = context or WorkbookWorkflowContext(workflow_kind="create_branch")
        base = Path(input_dir)
        for file_path in self._list_workbooks(base):
            rel_path = self._relative_path(base, file_path)
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    row_iter = sheet.iter_rows(values_only=True)
                    headers = list(next(row_iter, ()))
                    mapping = self._resolve_mapping(headers, project_id, context, strict=True)
                    for row_index, values in enumerate(row_iter, start=2):
                        yield self._extract_row(rel_path, sheet.title, row_index, values, mapping, context)
            finally:
                workbook.close()

    def _sample_issue_count(self, row_iter: Any, mapping: dict[str, Any], context: WorkbookWorkflowContext) -> int:
        issue_count = 0
        for index, values in enumerate(row_iter):
            if index >= self.SAMPLE_ROW_LIMIT:
                break
            row = self._extract_row("", "", index + 2, values, mapping, context)
            if row.status != "ok":
                issue_count += 1
        return issue_count

    def _resolve_mapping(
        self,
        headers: list[Any],
        project_id: int,
        context: WorkbookWorkflowContext,
        *,
        strict: bool,
    ) -> dict[str, Any]:
        schema = self.projects.get_schema(project_id)
        normalized = self._normalize_headers(headers)
        required_headers = [
            schema["fixed_columns"][field]
            for field in context.required_internal_fields()
        ]
        missing = [header for header in required_headers if header not in normalized]
        if strict and missing:
            raise ValueError(f"workbook missing required header: {missing[0]}")
        return {
            "available_headers": list(normalized.keys()),
            "missing_required_headers": missing,
            "file_name": normalized.get(ProjectService.FILE_NAME_HEADER),
            "business_key": normalized.get(schema["fixed_columns"]["business_key"]),
            "source": normalized.get(schema["fixed_columns"]["source"]),
            "translation_columns": {
                lang: normalized[lang]
                for lang in schema["translation_columns"]
                if lang in normalized
            },
            "remark_columns": {
                key: normalized[key]
                for key in schema["remark_columns"]
                if key in normalized
            },
        }

    def _extract_row(
        self,
        file_path: str,
        sheet_name: str,
        row_index: int,
        values: tuple[Any, ...],
        mapping: dict[str, Any],
        context: WorkbookWorkflowContext,
    ) -> WorkbookRow:
        business_key = self._cell(values, mapping.get("business_key"), is_content=False)
        source = self._cell(values, mapping.get("source"), is_content=False)
        file_name = None
        if mapping.get("file_name") is not None:
            file_name = self._cell(values, mapping.get("file_name"), is_content=False)
        translations = {
            lang: self._cell(values, index, is_content=True)
            for lang, index in mapping["translation_columns"].items()
        }
        remarks = {
            key: self._cell(values, index, is_content=False)
            for key, index in mapping["remark_columns"].items()
        }
        status = "ok"
        message = None
        if not business_key:
            status = "missing_business_key"
            message = "business_key is required"
        elif "source" in context.required_internal_fields() and not source:
            status = "missing_source"
            message = "source is required"
        return WorkbookRow(
            file_path=file_path,
            sheet_name=sheet_name,
            row_index=row_index,
            business_key=business_key,
            source=source,
            file_name=file_name,
            translations=translations,
            remarks=remarks,
            status=status,
            message=message,
        )

    def _cell(self, values: tuple[Any, ...], index: int | None, *, is_content: bool) -> str:
        if not index:
            return ""
        value = values[index - 1] if len(values) >= index else None
        return normalize_content_value(value) if is_content else normalize_non_content_value(value)

    def _normalize_headers(self, headers: list[Any]) -> dict[str, int]:
        normalized: dict[str, int] = {}
        for index, value in enumerate(headers):
            header = normalize_non_content_value(value)
            if header and header not in normalized:
                normalized[header] = index + 1
        return normalized

    def _list_workbooks(self, base: Path) -> list[Path]:
        if not base.exists() or not base.is_dir():
            raise ValueError(f"input directory not found: {base}")
        return [path for path in sorted(base.rglob("*.xlsx")) if not path.name.startswith("~$")]

    def _relative_path(self, base: Path, file_path: Path) -> str:
        return str(file_path.relative_to(base)).replace("\\", "/")
