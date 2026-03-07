from __future__ import annotations

import csv
import shutil
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from app.services.project_service import DEFAULT_PROJECT_ID, ProjectService
from app.services.string_service import StringService


class FillService:
    def __init__(self) -> None:
        self.projects = ProjectService()
        self.strings = StringService()

    def fill_and_export(
        self,
        source_dir: str,
        output_zip: str,
        lang: str,
        project_id: int = DEFAULT_PROJECT_ID,
        work_dir: str | None = None,
    ) -> dict[str, Any]:
        self.projects.require_language(lang, project_id)
        root = Path(source_dir)
        if not root.exists() or not root.is_dir():
            raise FileNotFoundError(f"fill source directory not found: {source_dir}")

        all_strings = self.strings.list_strings(project_id=project_id, include_deleted=False)
        strings_by_key = {item["business_key"]: item for item in all_strings}
        rel_keys = {
            item["business_key"]
            for item in self.strings.get_membership_strings("rel", "current", project_id)
        }

        export_dir = Path(work_dir) if work_dir else root.parent / f"{root.name}_filled"
        if export_dir.exists():
            shutil.rmtree(export_dir)
        export_dir.mkdir(parents=True, exist_ok=True)

        filled = miss = mismatch = kept = 0
        report_rows: list[dict[str, Any]] = []

        for path in sorted(root.rglob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            relative_path = path.relative_to(root)
            out_path = export_dir / relative_path
            out_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = load_workbook(path)
            for sheet in workbook.worksheets:
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                mapping = self.projects.resolve_headers(headers, project_id)
                target_column = mapping["translation_columns"][lang]
                for row_index in range(2, sheet.max_row + 1):
                    business_key = self._cell_text(sheet, row_index, mapping["business_key"])
                    if not business_key:
                        continue
                    source = self._cell_text(sheet, row_index, mapping["source"]) or ""
                    current_target = sheet.cell(row=row_index, column=target_column).value
                    candidate = strings_by_key.get(business_key)
                    if not candidate:
                        miss += 1
                        report_rows.append(
                            {
                                "file_path": str(relative_path).replace("\\", "/"),
                                "sheet_name": sheet.title,
                                "row_index": row_index,
                                "business_key": business_key,
                                "status": "MISSING_KEY_IN_BASE",
                            }
                        )
                        continue
                    if candidate["source"] != source:
                        mismatch += 1
                        report_rows.append(
                            {
                                "file_path": str(relative_path).replace("\\", "/"),
                                "sheet_name": sheet.title,
                                "row_index": row_index,
                                "business_key": business_key,
                                "status": "SRC_MISMATCH",
                            }
                        )
                        continue
                    if current_target not in (None, ""):
                        kept += 1
                    sheet.cell(row=row_index, column=target_column).value = candidate["translations"].get(lang)
                    filled += 1
                    report_rows.append(
                        {
                            "file_path": str(relative_path).replace("\\", "/"),
                            "sheet_name": sheet.title,
                            "row_index": row_index,
                            "business_key": business_key,
                            "status": "FILLED",
                            "from_scope": "rel" if business_key in rel_keys else "master",
                        }
                    )
            workbook.save(out_path)

        report_path = export_dir / "fill_report.csv"
        with report_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["file_path", "sheet_name", "row_index", "business_key", "status", "from_scope"])
            for row in report_rows:
                writer.writerow(
                    [
                        row.get("file_path"),
                        row.get("sheet_name"),
                        row.get("row_index"),
                        row.get("business_key"),
                        row.get("status"),
                        row.get("from_scope", ""),
                    ]
                )

        with ZipFile(output_zip, "w", compression=ZIP_DEFLATED) as archive:
            for file_path in sorted(export_dir.rglob("*")):
                if file_path.is_file():
                    archive.write(file_path, file_path.relative_to(export_dir))

        return {
            "filled_count": filled,
            "miss_key_count": miss,
            "src_mismatch_count": mismatch,
            "kept_original_count": kept,
            "report_path": str(report_path),
            "report_rows": report_rows,
            "output_zip": output_zip,
        }

    def _cell_text(self, sheet: Any, row_index: int, column_index: int | None) -> str | None:
        if not column_index:
            return None
        value = sheet.cell(row=row_index, column=column_index).value
        if value is None:
            return None
        text = str(value).strip()
        return text or None
