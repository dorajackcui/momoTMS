from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.db import get_conn, json_dumps, json_loads
from app.services.project_service import DEFAULT_PROJECT_ID, ProjectService
from app.services.utils import now_iso


class ImportService:
    def __init__(self) -> None:
        self.projects = ProjectService()

    def import_directory(
        self,
        input_dir: str,
        project_id: int = DEFAULT_PROJECT_ID,
    ) -> dict[str, Any]:
        base = Path(input_dir)
        if not base.exists() or not base.is_dir():
            raise FileNotFoundError(f"input directory not found: {input_dir}")
        files = [path for path in sorted(base.rglob("*.xlsx")) if not path.name.startswith("~$")]
        with get_conn() as conn:
            cur = conn.execute(
                """
                INSERT INTO imports(project_id, created_at, meta_json)
                VALUES (?, ?, ?)
                """,
                (
                    project_id,
                    now_iso(),
                    json_dumps({"input_dir": str(base), "project_id": project_id}),
                ),
            )
            batch_id = int(cur.lastrowid)

        rows_scanned = 0
        issues = 0
        for file_path in files:
            rel_path = str(file_path.relative_to(base)).replace("\\", "/")
            workbook = load_workbook(filename=file_path)
            for sheet in workbook.worksheets:
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                try:
                    mapping = self.projects.resolve_headers(headers, project_id)
                except ValueError as exc:
                    issues += 1
                    self._insert_row(
                        batch_id,
                        rel_path,
                        sheet.title,
                        0,
                        None,
                        None,
                        "sheet_error",
                        str(exc),
                        {},
                    )
                    continue
                for row_index in range(2, sheet.max_row + 1):
                    rows_scanned += 1
                    payload = self._extract_payload(sheet, row_index, mapping)
                    status = "ok"
                    message = None
                    if not payload["business_key"]:
                        status = "missing_business_key"
                        message = "business_key is required"
                        issues += 1
                    elif not payload["source"]:
                        status = "missing_source"
                        message = "source is required"
                        issues += 1
                    self._insert_row(
                        batch_id,
                        rel_path,
                        sheet.title,
                        row_index,
                        payload["business_key"] or None,
                        payload["source"] or None,
                        status,
                        message,
                        payload,
                    )

        summary = self.get_batch_summary(batch_id)
        summary["files_scanned"] = len(files)
        summary["rows_scanned"] = rows_scanned
        summary["issues"] = issues
        return summary

    def get_batch_summary(self, import_batch_id: int) -> dict[str, Any]:
        batches = {batch["import_batch_id"]: batch for batch in self.list_batches(limit=1000)}
        batch = batches.get(import_batch_id)
        if not batch:
            raise KeyError(f"import batch not found: {import_batch_id}")
        return batch

    def import_report(
        self,
        import_batch_id: int,
        issues_only: bool = False,
    ) -> dict[str, Any]:
        query = "SELECT * FROM import_rows WHERE import_batch_id = ?"
        params: list[Any] = [import_batch_id]
        if issues_only:
            query += " AND status != 'ok'"
        query += " ORDER BY import_row_id"
        with get_conn() as conn:
            rows = conn.execute(query, params).fetchall()
        summary = self.get_batch_summary(import_batch_id)
        report_rows = [
            {
                "import_row_id": int(row["import_row_id"]),
                "file_path": row["file_path"],
                "sheet_name": row["sheet_name"],
                "row_index": int(row["row_index"]),
                "business_key": row["business_key"],
                "source": row["source"],
                "status": row["status"],
                "message": row["message"],
                "payload": json_loads(row["payload_json"]),
            }
            for row in rows
        ]
        return {"summary": summary, "rows": report_rows}

    def list_batches(self, limit: int = 20) -> list[dict[str, Any]]:
        with get_conn() as conn:
            rows = conn.execute(
                """
                SELECT
                    i.import_batch_id,
                    i.created_at,
                    i.meta_json,
                    COUNT(ir.import_row_id) AS rows_scanned,
                    COUNT(DISTINCT ir.file_path) AS files_scanned,
                    SUM(CASE WHEN ir.status != 'ok' THEN 1 ELSE 0 END) AS issues
                FROM imports i
                LEFT JOIN import_rows ir ON ir.import_batch_id = i.import_batch_id
                GROUP BY i.import_batch_id, i.created_at, i.meta_json
                ORDER BY i.import_batch_id DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
        return [
            {
                "import_batch_id": int(row["import_batch_id"]),
                "created_at": row["created_at"],
                "meta": json_loads(row["meta_json"]),
                "rows_scanned": int(row["rows_scanned"] or 0),
                "files_scanned": int(row["files_scanned"] or 0),
                "issues": int(row["issues"] or 0),
            }
            for row in rows
        ]

    def _insert_row(
        self,
        import_batch_id: int,
        file_path: str,
        sheet_name: str,
        row_index: int,
        business_key: str | None,
        source: str | None,
        status: str,
        message: str | None,
        payload: dict[str, Any],
    ) -> None:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO import_rows(
                    import_batch_id,
                    file_path,
                    sheet_name,
                    row_index,
                    business_key,
                    source,
                    status,
                    message,
                    payload_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_batch_id,
                    file_path,
                    sheet_name,
                    row_index,
                    business_key,
                    source,
                    status,
                    message,
                    json_dumps(payload),
                ),
            )

    def _extract_payload(self, sheet: Any, row_index: int, mapping: dict[str, Any]) -> dict[str, Any]:
        def cell_text(column_index: int | None) -> str | None:
            if not column_index:
                return None
            value = sheet.cell(row=row_index, column=column_index).value
            if value is None:
                return None
            text = str(value).strip()
            return text or None

        translations = {
            lang: cell_text(column_index)
            for lang, column_index in mapping["translation_columns"].items()
        }
        remarks = {
            remark_key: cell_text(column_index)
            for remark_key, column_index in mapping["remark_columns"].items()
        }
        return {
            "file_name": cell_text(mapping["file_name"]),
            "business_key": cell_text(mapping["business_key"]),
            "source": cell_text(mapping["source"]),
            "translations": translations,
            "remarks": remarks,
        }
