from __future__ import annotations

from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import load_workbook

from app.db import get_conn, json_dumps, json_loads
from app.services.shared.io import normalize_content_value, normalize_non_content_value
from app.services.project.service import DEFAULT_PROJECT_ID, ProjectService
from app.services.shared.utils import now_iso


class ImportService:
    INSERT_CHUNK_SIZE = 1000

    def __init__(self) -> None:
        self.projects = ProjectService()

    def import_directory(
        self,
        input_dir: str,
        project_id: int = DEFAULT_PROJECT_ID,
        mapping_overrides: dict[str, dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        base = Path(input_dir)
        if not base.exists() or not base.is_dir():
            raise ValueError(f"input directory not found: {input_dir}")
        files = [path for path in sorted(base.rglob("*.xlsx")) if not path.name.startswith("~$")]
        started = perf_counter()
        with get_conn() as conn:
            cur = conn.execute(
                """
                INSERT INTO imports(project_id, created_at, meta_json)
                VALUES (?, ?, ?)
                """,
                (
                    project_id,
                    now_iso(),
                    json_dumps({"input_dir": str(base), "project_id": project_id}),
                ),
            )
            batch_id = int(cur.lastrowid)

        rows_scanned = 0
        issues = 0
        inserted_row_count = 0
        pending_rows: list[tuple[Any, ...]] = []
        persist_elapsed_ms = 0
        for file_path in files:
            rel_path = str(file_path.relative_to(base)).replace("\\", "/")
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    row_iter = sheet.iter_rows(values_only=True)
                    headers = list(next(row_iter, ()))
                    sheet_key = self.build_sheet_key(rel_path, sheet.title)
                    try:
                        mapping = self.projects.resolve_headers(
                            headers,
                            project_id,
                            override=(mapping_overrides or {}).get(sheet_key),
                        )
                    except ValueError as exc:
                        issues += 1
                        pending_rows.append(
                            self._build_import_row(
                                batch_id,
                                rel_path,
                                sheet.title,
                                0,
                                None,
                                None,
                                "sheet_error",
                                str(exc),
                                {},
                            )
                        )
                        inserted_row_count += len(pending_rows)
                        persist_elapsed_ms += self._flush_pending_rows(pending_rows)
                        continue
                    for row_index, row_values in enumerate(row_iter, start=2):
                        rows_scanned += 1
                        payload = self._extract_payload(row_values, mapping, rel_path)
                        status = "ok"
                        message = None
                        if not payload["business_key"]:
                            status = "missing_business_key"
                            message = "business_key is required"
                            issues += 1
                        elif not payload["source"]:
                            status = "missing_source"
                            message = "source is required"
                            issues += 1
                        pending_rows.append(
                            self._build_import_row(
                                batch_id,
                                rel_path,
                                sheet.title,
                                row_index,
                                payload["business_key"],
                                payload["source"],
                                status,
                                message,
                                payload,
                            )
                        )
                        if len(pending_rows) >= self.INSERT_CHUNK_SIZE:
                            inserted_row_count += len(pending_rows)
                            persist_elapsed_ms += self._flush_pending_rows(pending_rows)
            finally:
                workbook.close()

        inserted_row_count += len(pending_rows)
        persist_elapsed_ms += self._flush_pending_rows(pending_rows)
        total_elapsed_ms = int((perf_counter() - started) * 1000)
        parse_elapsed_ms = max(total_elapsed_ms - persist_elapsed_ms, 0)

        summary = self.get_batch_summary(batch_id)
        summary["files_scanned"] = len(files)
        summary["rows_scanned"] = rows_scanned
        summary["issues"] = issues
        summary["stages"] = [
            {
                "stage": "parse",
                "elapsed_ms": parse_elapsed_ms,
                "meta": {"files_scanned": len(files), "rows_scanned": rows_scanned},
            },
            {
                "stage": "persist_import",
                "elapsed_ms": persist_elapsed_ms,
                "meta": {"rows_inserted": inserted_row_count},
            },
        ]
        return summary

    def preview_directory(
        self,
        input_dir: str,
        project_id: int = DEFAULT_PROJECT_ID,
    ) -> dict[str, Any]:
        base = Path(input_dir)
        if not base.exists() or not base.is_dir():
            raise ValueError(f"input directory not found: {input_dir}")
        schema = self.projects.get_schema(project_id)
        sheet_previews: list[dict[str, Any]] = []
        unique_files: set[str] = set()
        files = [path for path in sorted(base.rglob("*.xlsx")) if not path.name.startswith("~$")]
        for file_path in files:
            clean_path = self._normalize_relative_path(str(file_path.relative_to(base)).replace("\\", "/"))
            unique_files.add(clean_path)
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    row_iter = sheet.iter_rows(values_only=True)
                    headers = list(next(row_iter, ()))
                    preview = self.projects.preview_headers(headers, project_id)
                    sheet_previews.append(
                        {
                            "sheet_key": self.build_sheet_key(clean_path, sheet.title),
                            "file_path": clean_path,
                            "derived_file_name": clean_path,
                            "sheet_name": sheet.title,
                            "available_headers": preview["available_headers"],
                            "suggested_mapping": preview["suggested_mapping"],
                            "missing_targets": preview["missing_targets"],
                            "auto_match_ready": preview["auto_match_ready"],
                        }
                    )
            finally:
                workbook.close()
        return {
            "schema": schema,
            "file_count": len(unique_files),
            "sheet_count": len(sheet_previews),
            "sheet_previews": sheet_previews,
        }

    def get_batch_summary(self, import_batch_id: int) -> dict[str, Any]:
        with get_conn() as conn:
            row = conn.execute(
                """
                SELECT
                    i.import_batch_id,
                    i.project_id,
                    i.created_at,
                    i.meta_json,
                    COUNT(ir.import_row_id) AS rows_scanned,
                    COUNT(DISTINCT ir.file_path) AS files_scanned,
                    SUM(CASE WHEN ir.status != 'ok' THEN 1 ELSE 0 END) AS issues
                FROM imports i
                LEFT JOIN import_rows ir ON ir.import_batch_id = i.import_batch_id
                WHERE i.import_batch_id = ?
                GROUP BY i.import_batch_id, i.project_id, i.created_at, i.meta_json
                """,
                (import_batch_id,),
            ).fetchone()
        if not row:
            raise KeyError(f"import batch not found: {import_batch_id}")
        return {
            "import_batch_id": int(row["import_batch_id"]),
            "project_id": int(row["project_id"]),
            "created_at": row["created_at"],
            "meta": json_loads(row["meta_json"]),
            "rows_scanned": int(row["rows_scanned"] or 0),
            "files_scanned": int(row["files_scanned"] or 0),
            "issues": int(row["issues"] or 0),
        }

    def import_report(
        self,
        import_batch_id: int,
        issues_only: bool = False,
        limit: int | None = None,
    ) -> dict[str, Any]:
        query = "SELECT * FROM import_rows WHERE import_batch_id = ?"
        params: list[Any] = [import_batch_id]
        if issues_only:
            query += " AND status != 'ok'"
        query += " ORDER BY import_row_id"
        if limit is not None:
            query += " LIMIT ?"
            params.append(limit)
        with get_conn() as conn:
            rows = conn.execute(query, params).fetchall()
        summary = self.get_batch_summary(import_batch_id)
        report_rows = [
            {
                "import_row_id": int(row["import_row_id"]),
                "file_path": row["file_path"],
                "sheet_name": row["sheet_name"],
                "row_index": int(row["row_index"]),
                "business_key": row["business_key"],
                "source": row["source"],
                "status": row["status"],
                "message": row["message"],
                "payload": json_loads(row["payload_json"]),
            }
            for row in rows
        ]
        return {"summary": summary, "rows": report_rows}

    def list_batches(
        self,
        limit: int = 20,
        project_id: int | None = None,
    ) -> list[dict[str, Any]]:
        params: list[Any] = []
        where = ""
        if project_id is not None:
            where = "WHERE i.project_id = ?"
            params.append(project_id)
        params.append(limit)
        with get_conn() as conn:
            rows = conn.execute(
                f"""
                SELECT
                    i.import_batch_id,
                    i.project_id,
                    i.created_at,
                    i.meta_json,
                    COUNT(ir.import_row_id) AS rows_scanned,
                    COUNT(DISTINCT ir.file_path) AS files_scanned,
                    SUM(CASE WHEN ir.status != 'ok' THEN 1 ELSE 0 END) AS issues
                FROM imports i
                LEFT JOIN import_rows ir ON ir.import_batch_id = i.import_batch_id
                {where}
                GROUP BY i.import_batch_id, i.project_id, i.created_at, i.meta_json
                ORDER BY i.import_batch_id DESC
                LIMIT ?
                """,
                params,
            ).fetchall()
        return [
            {
                "import_batch_id": int(row["import_batch_id"]),
                "project_id": int(row["project_id"]),
                "created_at": row["created_at"],
                "meta": json_loads(row["meta_json"]),
                "rows_scanned": int(row["rows_scanned"] or 0),
                "files_scanned": int(row["files_scanned"] or 0),
                "issues": int(row["issues"] or 0),
            }
            for row in rows
        ]

    def require_batch_project(self, import_batch_id: int, project_id: int) -> None:
        with get_conn() as conn:
            row = conn.execute(
                """
                SELECT project_id
                FROM imports
                WHERE import_batch_id = ?
                """,
                (import_batch_id,),
            ).fetchone()
        if not row:
            raise KeyError(f"import batch not found: {import_batch_id}")
        if int(row["project_id"]) != project_id:
            raise KeyError(f"import batch not found: {import_batch_id}")

    def _insert_rows(self, rows: list[tuple[Any, ...]]) -> None:
        if not rows:
            return
        with get_conn() as conn:
            conn.executemany(
                """
                INSERT INTO import_rows(
                    import_batch_id,
                    file_path,
                    sheet_name,
                    row_index,
                    business_key,
                    source,
                    status,
                    message,
                    payload_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows,
            )

    def _flush_pending_rows(self, rows: list[tuple[Any, ...]]) -> int:
        if not rows:
            return 0
        started = perf_counter()
        self._insert_rows(rows)
        rows.clear()
        return int((perf_counter() - started) * 1000)

    def _build_import_row(
        self,
        import_batch_id: int,
        file_path: str,
        sheet_name: str,
        row_index: int,
        business_key: str | None,
        source: str | None,
        status: str,
        message: str | None,
        payload: dict[str, Any],
    ) -> tuple[Any, ...]:
        return (
            import_batch_id,
            file_path,
            sheet_name,
            row_index,
            business_key,
            source,
            status,
            message,
            json_dumps(payload),
        )

    def _extract_payload(
        self,
        row_values: tuple[Any, ...],
        mapping: dict[str, Any],
        file_path: str,
    ) -> dict[str, Any]:
        def cell_value(column_index: int | None, *, is_content: bool) -> str:
            if not column_index:
                return ""
            value = row_values[column_index - 1] if len(row_values) >= column_index else None
            if is_content:
                return normalize_content_value(value)
            return normalize_non_content_value(value)

        translations = {
            lang: cell_value(column_index, is_content=True)
            for lang, column_index in mapping["translation_columns"].items()
        }
        remarks = {
            remark_key: cell_value(column_index, is_content=False)
            for remark_key, column_index in mapping["remark_columns"].items()
        }
        payload: dict[str, Any] = {
            "business_key": cell_value(mapping["business_key"], is_content=False),
            "source": cell_value(mapping["source"], is_content=False),
            "translations": translations,
            "remarks": remarks,
        }
        if mapping.get("file_name") is not None:
            payload["file_name"] = cell_value(mapping["file_name"], is_content=False)
        return payload

    @staticmethod
    def build_sheet_key(file_path: str, sheet_name: str) -> str:
        return f"{file_path}::{sheet_name}"

    @staticmethod
    def _normalize_relative_path(relative_path: str) -> str:
        cleaned = relative_path.replace("\\", "/").strip("/")
        if not cleaned or cleaned.startswith("../") or "/../" in f"/{cleaned}/":
            raise ValueError(f"invalid relative path: {relative_path}")
        if not cleaned.endswith(".xlsx"):
            raise ValueError(f"unsupported upload file: {relative_path}")
        return cleaned
