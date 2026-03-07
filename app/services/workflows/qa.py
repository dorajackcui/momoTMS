from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.shared.io import normalize_content_value, normalize_non_content_value
from app.services.project.service import DEFAULT_PROJECT_ID, ProjectService
from app.services.workflows.qa_rules import validate_pair


class QaScanService:
    def __init__(self) -> None:
        self.projects = ProjectService()

    def scan_directory(
        self,
        source_dir: str,
        lang: str,
        project_id: int = DEFAULT_PROJECT_ID,
    ) -> dict[str, Any]:
        self.projects.require_language(lang, project_id)
        root = Path(source_dir)
        if not root.exists() or not root.is_dir():
            raise FileNotFoundError(f"qa source directory not found: {source_dir}")

        scanned_rows = 0
        report_rows: list[dict[str, Any]] = []
        for path in sorted(root.rglob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            workbook = load_workbook(path)
            for sheet in workbook.worksheets:
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                mapping = self.projects.resolve_headers(headers, project_id)
                target_column = mapping["translation_columns"][lang]
                for row_index in range(2, sheet.max_row + 1):
                    business_key = self._cell_non_content(
                        sheet,
                        row_index,
                        mapping["business_key"],
                    )
                    if not business_key:
                        continue
                    scanned_rows += 1
                    source = self._cell_non_content(sheet, row_index, mapping["source"])
                    target = self._cell_content(sheet, row_index, target_column)
                    for result in validate_pair(source, target):
                        if result.ok:
                            continue
                        report_rows.append(
                            {
                                "file_path": str(path.relative_to(root)).replace("\\", "/"),
                                "sheet_name": sheet.title,
                                "row_index": row_index,
                                "business_key": business_key,
                                "lang": lang,
                                "rule": result.rule,
                                "src_excerpt": source[:120],
                                "tgt_excerpt": target[:120],
                            }
                        )

        rule_counts: dict[str, int] = {}
        for row in report_rows:
            rule = row["rule"]
            rule_counts[rule] = rule_counts.get(rule, 0) + 1
        return {
            "scanned_rows": scanned_rows,
            "issue_count": len(report_rows),
            "rule_counts": rule_counts,
            "report_rows": report_rows,
        }

    def _cell_non_content(self, sheet: Any, row_index: int, column_index: int | None) -> str:
        if not column_index:
            return ""
        value = sheet.cell(row=row_index, column=column_index).value
        return normalize_non_content_value(value)

    def _cell_content(self, sheet: Any, row_index: int, column_index: int | None) -> str:
        if not column_index:
            return ""
        value = sheet.cell(row=row_index, column=column_index).value
        return normalize_content_value(value)
