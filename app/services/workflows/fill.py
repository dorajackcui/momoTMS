from __future__ import annotations

import csv
import shutil
from pathlib import Path
from time import perf_counter
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from app.services.shared.io import (
    has_valid_fill_combined_key,
    is_blank_value,
    normalize_content_value,
    normalize_fill_combined_key,
    normalize_non_content_value,
)
from app.services.project.service import DEFAULT_PROJECT_ID, ProjectService
from app.services.variant.records import FillCandidateRecord
from app.services.workflows.fill_queries import FillQueryService


class FillService:
    def __init__(self) -> None:
        self.projects = ProjectService()
        self.fill_queries = FillQueryService()

    def fill_and_export(
        self,
        source_dir: str,
        output_zip: str,
        lang: str,
        project_id: int = DEFAULT_PROJECT_ID,
        work_dir: str | None = None,
        allow_blank_write: bool = False,
    ) -> dict[str, Any]:
        self.projects.require_language(lang, project_id)
        root = Path(source_dir)
        if not root.exists() or not root.is_dir():
            raise ValueError(f"fill source directory not found: {source_dir}")

        candidates = self.fill_queries.list_fill_candidates(project_id, lang)
        keys_in_project, strings_by_combo = self._build_fill_indexes(candidates)

        export_dir = Path(work_dir) if work_dir else root.parent / f"{root.name}_filled"
        if export_dir.exists():
            shutil.rmtree(export_dir)
        export_dir.mkdir(parents=True, exist_ok=True)

        filled = miss = mismatch = kept = skipped_invalid = skipped_blank = 0
        report_rows: list[dict[str, Any]] = []
        fill_started = perf_counter()

        for path in sorted(root.rglob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            relative_path = path.relative_to(root)
            out_path = export_dir / relative_path
            out_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = load_workbook(path)
            for sheet in workbook.worksheets:
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                mapping = self.projects.resolve_headers(headers, project_id)
                target_column = mapping["translation_columns"].get(lang)
                if target_column is None:
                    raise ValueError(f"workbook missing required header: {lang}")
                for row_index in range(2, sheet.max_row + 1):
                    business_key = self._cell_non_content(
                        sheet,
                        row_index,
                        mapping["business_key"],
                    )
                    source = self._cell_non_content(sheet, row_index, mapping["source"])
                    if not has_valid_fill_combined_key(business_key, source):
                        skipped_invalid += 1
                        report_rows.append(
                            {
                                "file_path": str(relative_path).replace("\\", "/"),
                                "sheet_name": sheet.title,
                                "row_index": row_index,
                                "business_key": business_key,
                                "status": "SKIPPED_INVALID_COMBINED_KEY",
                            }
                        )
                        continue
                    combined_key = normalize_fill_combined_key(business_key, source)
                    candidate = strings_by_combo.get(combined_key)
                    if candidate is not None:
                        current_target = self._cell_content(sheet, row_index, target_column)
                        if not is_blank_value(current_target):
                            kept += 1
                        candidate_content = candidate["target_text"]
                        candidate_state = self._candidate_state(candidate)
                        if is_blank_value(candidate_content) and not allow_blank_write:
                            skipped_blank += 1
                            report_rows.append(
                                {
                                    "file_path": str(relative_path).replace("\\", "/"),
                                    "sheet_name": sheet.title,
                                    "row_index": row_index,
                                    "business_key": business_key,
                                    "status": "SKIPPED_BLANK_CONTENT",
                                    "match_variant_id": candidate["variant_id"],
                                    "match_variant_state": candidate_state,
                                }
                            )
                            continue
                        sheet.cell(row=row_index, column=target_column).value = candidate_content
                        filled += 1
                        report_rows.append(
                            {
                                "file_path": str(relative_path).replace("\\", "/"),
                                "sheet_name": sheet.title,
                                "row_index": row_index,
                                "business_key": business_key,
                                "status": "FILLED",
                                "match_variant_id": candidate["variant_id"],
                                "match_variant_state": candidate_state,
                            }
                        )
                        continue
                    if business_key not in keys_in_project:
                        miss += 1
                        report_rows.append(
                            {
                                "file_path": str(relative_path).replace("\\", "/"),
                                "sheet_name": sheet.title,
                                "row_index": row_index,
                                "business_key": business_key,
                                "status": "MISSING_KEY_IN_PROJECT",
                            }
                        )
                        continue
                    mismatch += 1
                    report_rows.append(
                        {
                            "file_path": str(relative_path).replace("\\", "/"),
                            "sheet_name": sheet.title,
                            "row_index": row_index,
                            "business_key": business_key,
                            "status": "SRC_MISMATCH",
                        }
                    )
            workbook.save(out_path)

        fill_elapsed_ms = int((perf_counter() - fill_started) * 1000)
        artifact_started = perf_counter()
        report_path = export_dir / "fill_report.csv"
        with report_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "file_path",
                    "sheet_name",
                    "row_index",
                    "business_key",
                    "status",
                    "match_variant_id",
                    "match_variant_state",
                ]
            )
            for row in report_rows:
                writer.writerow(
                    [
                        row.get("file_path"),
                        row.get("sheet_name"),
                        row.get("row_index"),
                        row.get("business_key"),
                        row.get("status"),
                        row.get("match_variant_id", ""),
                        row.get("match_variant_state", ""),
                    ]
                )

        with ZipFile(output_zip, "w", compression=ZIP_DEFLATED) as archive:
            for file_path in sorted(export_dir.rglob("*")):
                if file_path.is_file():
                    archive.write(file_path, file_path.relative_to(export_dir))
        artifact_elapsed_ms = int((perf_counter() - artifact_started) * 1000)

        return {
            "filled_count": filled,
            "miss_key_count": miss,
            "src_mismatch_count": mismatch,
            "kept_original_count": kept,
            "skipped_invalid_combined_key_count": skipped_invalid,
            "skipped_blank_content_count": skipped_blank,
            "report_path": str(report_path),
            "report_rows": report_rows,
            "output_zip": output_zip,
            "stages": [
                {
                    "stage": "fill_export",
                    "elapsed_ms": fill_elapsed_ms,
                    "meta": {"filled_count": filled, "row_count": len(report_rows)},
                },
                {
                    "stage": "artifact_write",
                    "elapsed_ms": artifact_elapsed_ms,
                    "meta": {"artifact_name": Path(output_zip).name},
                },
            ],
        }

    def _cell_non_content(self, sheet: Any, row_index: int, column_index: int | None) -> str:
        if not column_index:
            return ""
        value = sheet.cell(row=row_index, column=column_index).value
        return normalize_non_content_value(value)

    def _cell_content(self, sheet: Any, row_index: int, column_index: int | None) -> str:
        if not column_index:
            return ""
        value = sheet.cell(row=row_index, column=column_index).value
        return normalize_content_value(value)

    def _build_fill_indexes(
        self,
        candidates: list[FillCandidateRecord],
    ) -> tuple[set[str], dict[tuple[str, str], FillCandidateRecord]]:
        keys_in_project: set[str] = set()
        candidates_by_combo: dict[tuple[str, str], list[FillCandidateRecord]] = {}
        for candidate in candidates:
            business_key = normalize_non_content_value(candidate["business_key"])
            source = normalize_non_content_value(candidate["source"])
            keys_in_project.add(business_key)
            candidates_by_combo.setdefault((business_key, source), []).append(candidate)

        best_candidates: dict[tuple[str, str], FillCandidateRecord] = {}
        for combined_key, grouped_candidates in candidates_by_combo.items():
            best_candidates[combined_key] = self._pick_best_candidate(combined_key, grouped_candidates)
        return keys_in_project, best_candidates

    def _pick_best_candidate(
        self,
        combined_key: tuple[str, str],
        candidates: list[FillCandidateRecord],
    ) -> FillCandidateRecord:
        live_candidates = [candidate for candidate in candidates if candidate["trashed_at"] is None]
        if len(live_candidates) > 1:
            raise RuntimeError(
                "duplicate non-trashed fill candidates found for "
                f"business_key={combined_key[0]!r}, source={combined_key[1]!r}"
            )
        if live_candidates:
            return live_candidates[0]
        return sorted(
            candidates,
            key=lambda candidate: (candidate["updated_at"], candidate["variant_id"]),
            reverse=True,
        )[0]

    def _candidate_state(self, candidate: FillCandidateRecord) -> str:
        if candidate["trashed_at"] is not None:
            return "trashed"
        if candidate["orphaned_at"] is not None:
            return "orphan"
        return "active"
