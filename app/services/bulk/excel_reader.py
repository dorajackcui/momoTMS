from __future__ import annotations

from pathlib import Path
from typing import Any, Iterator

import openpyxl

from app.services.shared.io import (
    is_blank_value,
    normalize_content_map,
    normalize_non_content_value,
)


class BulkSeedError(Exception):
    def __init__(self, message: str, *, file_name: str = "", sheet_name: str = "", row_index: int = 0) -> None:
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.row_index = row_index
        super().__init__(message)


def read_excel_chunks(
    workbook_path: str,
    schema: dict[str, Any],
    chunk_size: int = 5000,
) -> Iterator[list[dict[str, Any]]]:
    fixed_columns = schema["fixed_columns"]
    bk_header = fixed_columns["business_key"]
    src_header = fixed_columns["source"]
    translation_cols = schema["translation_columns"]
    remark_cols = schema["remark_columns"]

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            sheet_name = ws.title
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row is None:
                continue
            headers = [normalize_non_content_value(h) for h in header_row]
            col_map = _build_column_map(
                headers,
                bk_header=bk_header,
                src_header=src_header,
                translation_cols=translation_cols,
                remark_cols=remark_cols,
                file_name=Path(workbook_path).name,
                sheet_name=sheet_name,
            )
            file_name = Path(workbook_path).name
            chunk: list[dict[str, Any]] = []
            for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                parsed = _parse_row(
                    row,
                    col_map=col_map,
                    file_name=file_name,
                    sheet_name=sheet_name,
                    row_index=row_index,
                    translation_cols=translation_cols,
                    remark_cols=remark_cols,
                )
                chunk.append(parsed)
                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []
            if chunk:
                yield chunk
    finally:
        wb.close()


def _build_column_map(
    headers: list[str],
    *,
    bk_header: str,
    src_header: str,
    translation_cols: list[str],
    remark_cols: list[str],
    file_name: str,
    sheet_name: str,
) -> dict[str, int]:
    header_index = {h: i for i, h in enumerate(headers) if h}
    col_map: dict[str, int] = {}
    if bk_header not in header_index:
        raise BulkSeedError(
            f"missing required header: {bk_header}",
            file_name=file_name,
            sheet_name=sheet_name,
        )
    col_map["business_key"] = header_index[bk_header]
    if src_header not in header_index:
        raise BulkSeedError(
            f"missing required header: {src_header}",
            file_name=file_name,
            sheet_name=sheet_name,
        )
    col_map["source"] = header_index[src_header]
    for lang in translation_cols:
        if lang in header_index:
            col_map[f"t:{lang}"] = header_index[lang]
    for remark_key in remark_cols:
        if remark_key in header_index:
            col_map[f"r:{remark_key}"] = header_index[remark_key]
    return col_map


def _parse_row(
    row: tuple[Any, ...],
    *,
    col_map: dict[str, int],
    file_name: str,
    sheet_name: str,
    row_index: int,
    translation_cols: list[str],
    remark_cols: list[str],
) -> dict[str, Any]:
    def cell(key: str) -> Any:
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    business_key = normalize_non_content_value(cell("business_key"))
    source = normalize_non_content_value(cell("source"))
    if is_blank_value(business_key):
        raise BulkSeedError(
            f"blank business_key at row {row_index}",
            file_name=file_name,
            sheet_name=sheet_name,
            row_index=row_index,
        )
    if is_blank_value(source):
        raise BulkSeedError(
            f"blank source at row {row_index}",
            file_name=file_name,
            sheet_name=sheet_name,
            row_index=row_index,
        )
    translations: dict[str, Any] = {}
    for lang in translation_cols:
        key = f"t:{lang}"
        if key in col_map:
            translations[lang] = cell(key)
    remarks: dict[str, Any] = {}
    for remark_key in remark_cols:
        key = f"r:{remark_key}"
        if key in col_map:
            remarks[remark_key] = cell(key)
    return {
        "business_key": business_key,
        "source": source,
        "file_name": file_name,
        "sheet_name": sheet_name,
        "row_index": row_index,
        "translations": normalize_content_map(translations),
        "remarks": {k: normalize_non_content_value(v) for k, v in remarks.items()},
    }
