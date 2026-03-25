from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.db import get_conn, get_db_path, init_db
from app.services.branch.models import BranchRef
from app.services.demo.service import DemoService
from app.services.project.service import ProjectService
from app.services.variant.catalog import VariantCatalogService
from app.services.variant.entries import EntryService
from app.services.workflows.fill import FillService
from app.services.workflows.trash_restore import TrashRestoreService
from tests.service_helpers import branch_services


def reset_demo() -> dict:
    db_path = get_db_path()
    if Path(db_path).exists():
        Path(db_path).unlink()
    DemoService().reset()
    return DemoService().get_sample("core-cycle")


def write_fill_workbook(root: Path, relative_path: str, rows: list[list[object]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in rows:
        sheet.append(row)
    output_path = root / relative_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def read_target_text(workbook_path: Path, row_index: int = 2, column_index: int = 4) -> str | None:
    workbook = load_workbook(workbook_path)
    return workbook.active.cell(row=row_index, column=column_index).value


def output_workbook_path(source_dir: Path, relative_path: str) -> Path:
    return source_dir.parent / f"{source_dir.name}_filled" / relative_path


def report_row_by_key(report_rows: list[dict], business_key: str) -> dict:
    for row in report_rows:
        if row.get("business_key") == business_key:
            return row
    raise KeyError(f"missing report row for business key: {business_key}")


def touch_variant_updated_at(variant_id: int, updated_at: str) -> None:
    with get_conn() as conn:
        conn.execute(
            "UPDATE variants SET updated_at = ? WHERE variant_id = ?",
            (updated_at, variant_id),
        )


def test_fill_reads_project_variants_across_active_orphan_and_missing_rows() -> None:
    sample = reset_demo()
    source_dir = Path(sample["paths"]["fill_dir"])
    output_zip = Path(sample["paths"]["root"]) / "filled.zip"
    result = FillService().fill_and_export(
        str(source_dir),
        str(output_zip),
        sample["lang"],
    )

    active_row = report_row_by_key(result["report_rows"], "common.welcome")
    orphan_row = report_row_by_key(result["report_rows"], "fill.master_only")
    mismatch_row = report_row_by_key(result["report_rows"], "fill.rel")
    missing_row = report_row_by_key(result["report_rows"], "fill.missing")

    assert active_row["status"] == "FILLED"
    assert active_row["match_variant_state"] == "active"
    assert active_row["pivot_lang"] is None
    assert active_row["pivot_sync_status"] is None
    assert orphan_row["status"] == "FILLED"
    assert orphan_row["match_variant_state"] == "orphan"
    assert orphan_row["pivot_lang"] is None
    assert orphan_row["pivot_sync_status"] is None
    assert mismatch_row["status"] == "SRC_MISMATCH"
    assert mismatch_row["pivot_lang"] is None
    assert mismatch_row["pivot_sync_status"] is None
    assert missing_row["status"] == "MISSING_KEY_IN_PROJECT"
    assert missing_row["pivot_lang"] is None
    assert missing_row["pivot_sync_status"] is None
    assert output_zip.exists()

    filled_workbook = output_workbook_path(source_dir, "fill/fill_input.xlsx")
    assert read_target_text(filled_workbook, row_index=2) == "Bienvenue {0}"
    assert read_target_text(filled_workbook, row_index=3) == "Depuis master"


def test_fill_uses_trashed_candidate_when_no_live_variant_exists() -> None:
    sample = reset_demo()
    read_service = branch_services()
    variant_service = TrashRestoreService()
    entry = read_service.entries.get_entry("trash.me")
    assert entry is not None
    original_variant = read_service.catalog.list_variants(int(entry["entry_id"]), include_trashed=True)[0]
    read_service.bindings.bind_scope(int(entry["entry_id"]), BranchRef.rel_current(), int(original_variant["variant_id"]))
    variant_service.delete(BranchRef.rel_current(), ["trash.me"])

    source_dir = Path(sample["paths"]["root"]) / "trashed-only-source"
    write_fill_workbook(
        source_dir,
        "single.xlsx",
        [
            ["file_name", "business_key", "source", "fr", "en", "context"],
            ["trash.xlsx", "trash.me", "Trash me source", "", "", "trashed-only"],
        ],
    )

    result = FillService().fill_and_export(
        str(source_dir),
        str(Path(sample["paths"]["root"]) / "trashed-only.zip"),
        sample["lang"],
    )

    row = report_row_by_key(result["report_rows"], "trash.me")
    assert row["status"] == "FILLED"
    assert row["match_variant_id"] == int(original_variant["variant_id"])
    assert row["match_variant_state"] == "trashed"
    assert row["pivot_lang"] is None
    assert row["pivot_sync_status"] is None
    assert read_target_text(output_workbook_path(source_dir, "single.xlsx")) == "Supprimer moi"


def test_fill_prefers_live_candidate_over_trashed_history() -> None:
    sample = reset_demo()
    read_service = branch_services()
    variant_service = TrashRestoreService()
    entry = read_service.entries.get_entry("trash.me")
    assert entry is not None
    entry_id = int(entry["entry_id"])
    original_variant = read_service.catalog.list_variants(entry_id, include_trashed=True)[0]
    read_service.bindings.bind_scope(entry_id, BranchRef.rel_current(), int(original_variant["variant_id"]))
    variant_service.delete(BranchRef.rel_current(), ["trash.me"])

    live_variant_id = read_service.catalog.create_variant(
        entry_id,
        read_service.catalog.build_content(
            "trash-live.xlsx",
            "Trash me source",
            {"fr": "Live translation wins", "en": "Live translation wins"},
            {"context": "live replacement"},
        ),
    )
    read_service.bindings.bind_scope(entry_id, BranchRef.rel_current(), live_variant_id)

    source_dir = Path(sample["paths"]["root"]) / "prefer-live-source"
    write_fill_workbook(
        source_dir,
        "single.xlsx",
        [
            ["file_name", "business_key", "source", "fr", "en", "context"],
            ["trash.xlsx", "trash.me", "Trash me source", "", "", "prefer-live"],
        ],
    )

    result = FillService().fill_and_export(
        str(source_dir),
        str(Path(sample["paths"]["root"]) / "prefer-live.zip"),
        sample["lang"],
    )

    row = report_row_by_key(result["report_rows"], "trash.me")
    assert row["status"] == "FILLED"
    assert row["match_variant_id"] == live_variant_id
    assert row["match_variant_state"] == "active"
    assert row["pivot_lang"] is None
    assert row["pivot_sync_status"] is None
    assert read_target_text(output_workbook_path(source_dir, "single.xlsx")) == "Live translation wins"


def test_fill_uses_latest_trashed_candidate_when_only_trashed_history_exists() -> None:
    sample = reset_demo()
    read_service = branch_services()
    variant_service = TrashRestoreService()
    entry = read_service.entries.get_entry("trash.me")
    assert entry is not None
    entry_id = int(entry["entry_id"])
    first_variant = read_service.catalog.list_variants(entry_id, include_trashed=True)[0]
    read_service.bindings.bind_scope(entry_id, BranchRef.rel_current(), int(first_variant["variant_id"]))
    variant_service.delete(BranchRef.rel_current(), ["trash.me"])

    second_variant_id = read_service.catalog.create_variant(
        entry_id,
        read_service.catalog.build_content(
            "trash-newer.xlsx",
            "Trash me source",
            {"fr": "Newest trashed translation", "en": "Newest trashed translation"},
            {"context": "newest trashed"},
        ),
    )
    read_service.bindings.bind_scope(entry_id, BranchRef.rel_current(), second_variant_id)
    variant_service.delete(BranchRef.rel_current(), ["trash.me"])
    touch_variant_updated_at(int(first_variant["variant_id"]), "2024-01-01T00:00:00+00:00")
    touch_variant_updated_at(second_variant_id, "2024-01-02T00:00:00+00:00")

    source_dir = Path(sample["paths"]["root"]) / "latest-trashed-source"
    write_fill_workbook(
        source_dir,
        "single.xlsx",
        [
            ["file_name", "business_key", "source", "fr", "en", "context"],
            ["trash.xlsx", "trash.me", "Trash me source", "", "", "latest-trashed"],
        ],
    )

    result = FillService().fill_and_export(
        str(source_dir),
        str(Path(sample["paths"]["root"]) / "latest-trashed.zip"),
        sample["lang"],
    )

    row = report_row_by_key(result["report_rows"], "trash.me")
    assert row["status"] == "FILLED"
    assert row["match_variant_id"] == second_variant_id
    assert row["match_variant_state"] == "trashed"
    assert row["pivot_lang"] is None
    assert row["pivot_sync_status"] is None
    assert read_target_text(output_workbook_path(source_dir, "single.xlsx")) == "Newest trashed translation"


def test_fill_reports_pivot_status_for_matched_variants_only() -> None:
    db_path = get_db_path()
    if Path(db_path).exists():
        Path(db_path).unlink()
    init_db()

    project = ProjectService().create_project(
        "Pivot Fill Project",
        ["fr", "en"],
        ["context"],
        {"fr": "en"},
    )
    project_id = int(project["project_id"])
    entries = EntryService()
    catalog = VariantCatalogService()
    entry = entries.get_or_create_entry("pivot.fill", project_id=project_id)
    variant_id = catalog.create_variant(
        int(entry["entry_id"]),
        catalog.build_content(
            "pivot-fill.xlsx",
            "Hello",
            {"en": "Hello", "fr": "Bonjour"},
            {"context": "pivot fill"},
        ),
    )
    catalog.update_variant(
        variant_id,
        catalog.build_content(
            "pivot-fill.xlsx",
            "Hello",
            {"en": "Hello there", "fr": "Bonjour"},
            {"context": "pivot fill"},
        ),
    )

    source_dir = Path(db_path).parent / "pivot-fill-source"
    output_zip = source_dir.parent / "pivot-fill.zip"
    write_fill_workbook(
        source_dir,
        "single.xlsx",
        [
            ["file_name", "business_key", "source", "fr", "en", "context"],
            ["pivot.xlsx", "pivot.fill", "Hello", "", "", "matched"],
            ["pivot.xlsx", "pivot.fill", "Hello mismatch", "", "", "mismatch"],
            ["pivot.xlsx", "pivot.missing", "Missing", "", "", "missing"],
        ],
    )

    result = FillService().fill_and_export(str(source_dir), str(output_zip), "fr", project_id=project_id)

    matched_row = report_row_by_key(result["report_rows"], "pivot.fill")
    mismatch_row = next(row for row in result["report_rows"] if row["status"] == "SRC_MISMATCH")
    missing_row = report_row_by_key(result["report_rows"], "pivot.missing")

    assert matched_row["status"] == "FILLED"
    assert matched_row["pivot_lang"] == "en"
    assert matched_row["pivot_sync_status"] == "PIVOT_OUT_OF_SYNC"
    assert mismatch_row["pivot_lang"] == "en"
    assert mismatch_row["pivot_sync_status"] is None
    assert missing_row["pivot_lang"] == "en"
    assert missing_row["pivot_sync_status"] is None
    assert read_target_text(output_workbook_path(source_dir, "single.xlsx")) == "Bonjour"
