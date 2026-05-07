from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.demo.service import DemoService
from app.services.project.service import ProjectService
from app.services.workbooks.models import WorkbookWorkflowContext
from app.services.workbooks.parser import WorkbookParser


def reset_demo() -> None:
    DemoService().reset()


def write_workbook(root: Path, relative_path: str, rows: list[list[object]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in rows:
        sheet.append(row)
    output = root / relative_path
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    return output


def test_parser_uses_project_workbook_headers(tmp_path) -> None:
    reset_demo()
    project = ProjectService().create_project(
        "Parser Contract Project",
        ["fr"],
        ["context"],
        business_key_header="key",
        source_header="source_text",
    )
    root = tmp_path / "input"
    write_workbook(
        root,
        "bundle/messages.xlsx",
        [
            ["key", "source_text", "fr", "context"],
            ["hello.key", "Hello", "Bonjour", "Greeting"],
        ],
    )

    context = WorkbookWorkflowContext(workflow_kind="create_branch")
    preview = WorkbookParser().precheck_directory(root, int(project["project_id"]), context)
    rows = list(WorkbookParser().iter_rows(root, int(project["project_id"]), context))

    assert preview.missing_required_headers == []
    assert preview.file_count == 1
    assert preview.sheet_count == 1
    assert rows[0].business_key == "hello.key"
    assert rows[0].source == "Hello"
    assert rows[0].translations == {"fr": "Bonjour"}
    assert rows[0].remarks == {"context": "Greeting"}


def test_trash_parser_requires_key_only(tmp_path) -> None:
    reset_demo()
    project = ProjectService().create_project(
        "Trash Parser Project",
        ["fr"],
        ["context"],
        business_key_header="key",
        source_header="source_text",
    )
    root = tmp_path / "trash"
    write_workbook(root, "trash.xlsx", [["key"], ["obsolete.key"]])

    context = WorkbookWorkflowContext(workflow_kind="branch_trash")
    preview = WorkbookParser().precheck_directory(root, int(project["project_id"]), context)
    rows = list(WorkbookParser().iter_rows(root, int(project["project_id"]), context))

    assert preview.missing_required_headers == []
    assert rows[0].business_key == "obsolete.key"
    assert rows[0].source == ""


def test_content_mutation_parser_requires_source(tmp_path) -> None:
    reset_demo()
    project = ProjectService().create_project(
        "Content Parser Project",
        ["fr"],
        ["context"],
        business_key_header="key",
        source_header="source_text",
    )
    root = tmp_path / "content"
    write_workbook(root, "content.xlsx", [["key", "fr"], ["hello.key", "Bonjour"]])

    context = WorkbookWorkflowContext(workflow_kind="branch_mutation", mutation_type="content")
    preview = WorkbookParser().precheck_directory(root, int(project["project_id"]), context)

    assert preview.missing_required_headers == ["source_text"]


from app.services.workbooks.batches import WorkbookBatchService


def test_workbook_batch_service_persists_rows_for_batch_reader(tmp_path) -> None:
    reset_demo()
    project = ProjectService().create_project(
        "Batch Project",
        ["fr"],
        ["context"],
        business_key_header="key",
        source_header="source_text",
    )
    root = tmp_path / "batch"
    write_workbook(
        root,
        "bundle/messages.xlsx",
        [
            ["key", "source_text", "fr", "context"],
            ["hello.key", "Hello", "Bonjour", "Greeting"],
        ],
    )

    context = WorkbookWorkflowContext(workflow_kind="branch_mutation", mutation_type="range")
    batch = WorkbookBatchService().create_batch_from_directory(root, int(project["project_id"]), context)
    rows = list(WorkbookBatchService().iter_rows(batch["workbook_batch_id"], int(project["project_id"])))

    assert batch["workbook_batch_id"] > 0
    assert batch["rows_scanned"] == 1
    assert batch["issues"] == 0
    assert rows[0]["business_key"] == "hello.key"
    assert rows[0]["source"] == "Hello"
    assert rows[0]["payload"]["translations"] == {"fr": "Bonjour"}


def test_workbook_batch_reads_file_name_from_source_name_column(tmp_path) -> None:
    reset_demo()
    project = ProjectService().create_project(
        "Batch Source Name Project",
        ["fr"],
        ["context"],
        business_key_header="key",
        source_header="source_text",
    )
    root = tmp_path / "batch-source-name"
    write_workbook(
        root,
        "bundle/messages.xlsx",
        [
            ["Source.Name", "key", "source_text", "fr"],
            ["business/from-column.xlsx", "hello.key", "Hello", "Bonjour"],
        ],
    )

    batch = WorkbookBatchService().create_batch_from_directory(
        root,
        int(project["project_id"]),
        WorkbookWorkflowContext(workflow_kind="branch_mutation", mutation_type="range"),
    )
    rows = list(WorkbookBatchService().iter_rows(batch["workbook_batch_id"], int(project["project_id"])))

    assert rows[0]["file_path"] == "bundle/messages.xlsx"
    assert rows[0]["payload"]["file_name"] == "business/from-column.xlsx"


def test_workbook_batch_service_iter_row_chunks_preserves_order(tmp_path) -> None:
    reset_demo()
    project = ProjectService().create_project(
        "Chunked Batch Project",
        ["fr"],
        ["context"],
        business_key_header="Key",
        source_header="MsgStr",
    )
    project_id = int(project["project_id"])
    root = tmp_path / "batch-chunks"
    write_workbook(
        root,
        "batch.xlsx",
        [
            ["Key", "MsgStr", "fr", "context"],
            ["chunk.1", "Source 1", "FR 1", "Remark 1"],
            ["chunk.2", "Source 2", "FR 2", "Remark 2"],
            ["chunk.3", "Source 3", "FR 3", "Remark 3"],
        ],
    )
    batch = WorkbookBatchService().create_batch_from_directory(
        root,
        project_id,
        WorkbookWorkflowContext(workflow_kind="branch_mutation", mutation_type="content"),
    )
    chunks = list(
        WorkbookBatchService().iter_row_chunks(
            batch["workbook_batch_id"],
            project_id,
            ok_only=True,
            chunk_size=2,
        )
    )
    assert [[row["business_key"] for row in chunk] for chunk in chunks] == [
        ["chunk.1", "chunk.2"],
        ["chunk.3"],
    ]
    iter_rows_keys = [
        row["business_key"]
        for row in WorkbookBatchService().iter_rows(batch["workbook_batch_id"], project_id, ok_only=True)
    ]
    assert [row["business_key"] for chunk in chunks for row in chunk] == iter_rows_keys


def test_workbook_batch_service_iter_row_chunks_rejects_invalid_chunk_size(tmp_path) -> None:
    reset_demo()
    project = ProjectService().create_project(
        "Invalid Chunked Batch Project",
        ["fr"],
        ["context"],
        business_key_header="Key",
        source_header="MsgStr",
    )
    project_id = int(project["project_id"])
    root = tmp_path / "batch-invalid-chunks"
    write_workbook(
        root,
        "batch.xlsx",
        [
            ["Key", "MsgStr", "fr", "context"],
            ["chunk.1", "Source 1", "FR 1", "Remark 1"],
        ],
    )
    batch = WorkbookBatchService().create_batch_from_directory(
        root,
        project_id,
        WorkbookWorkflowContext(workflow_kind="branch_mutation", mutation_type="content"),
    )

    with pytest.raises(ValueError, match="chunk_size"):
        list(
            WorkbookBatchService().iter_row_chunks(
                batch["workbook_batch_id"],
                project_id,
                chunk_size=0,
            )
        )
